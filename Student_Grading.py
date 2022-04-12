from openpyxl import Workbook
book=Workbook()
sheet=book.active
sheet['A1'] = "Name"
sheet['B1'] = "Overall Grade"
sheet['C1'] = "Maths"
sheet['D1'] = "Physics"
sheet['E1'] = "Chemistry"
sheet['F1'] = "Language"
i=1 #Coloumns A,B,C
j=1 #Rows
student=int(input("Enter no.of students in the class :"))
for j in range(student):
    sheet.cell(row=j+2, column=i).value = input("Enter the name of the student :")
    sheet.cell(row=j+2, column=i+1).value = input("Enter students overall grade:")
    sheet.cell(row=j+2, column=i+2).value = input("Enter students grade in Maths:")
    sheet.cell(row=j+2, column=i+3).value = input("Enter students grade in Physics:")
    sheet.cell(row=j+2, column=i+4).value = input("Enter students grade in Chemistry:")
    sheet.cell(row=j+2, column=i+5).value = input("Enter students grade in Language:")
    j=j+1

book.save(filename="Student.xlsx")