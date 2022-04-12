import openpyxl
path = ("/Users/sakethmaddineni/PycharmProjects/project_001/Toyota.xlsx")
wb_file = openpyxl.load_workbook(path)
wb_file
sheet = wb_file.active
sheet
sheet.cell(row=5,colunm=2)
val1 = sheet.cell(row=5,column=2)
val1
sheet.max_row
sheet.max_column
