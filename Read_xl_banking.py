import openpyxl
path = ("/Users/sakethmaddineni/PycharmProjects/project_001/bank_marketing.xlsx")
wb=openpyxl.load_workbook(path)
sheet=wb.active
val1=sheet.cell(row=4,column=5)
val=val1.value
print(val)

for i in range(1,51):
    for j in range(1,6):
        print(sheet.cell(row=i,column=j).value)
#Merge and center the cells F1 to N1
sheet.merge_cells("F1:N1")
sheet.delete_cols(idx=7,amount=11)
sheet.delete_rows(idx=100,amount=sheet.max_row)
sheet.merge_cells("H2:N22")
wb.save(filename="class1.xlsx")

"""wb_file=openpyxl.load_workbook("test1.xlsx")
sheet=wb_file.active
sheet.delete_cols(idx=6,amount=200)
wb_file.save(filename="test1.xlsx")
max_row=sheet.max_row
#print(max_row)
max_col=sheet.max_column
#print(max_col)
cells=sheet['A1':'E25']
for c1,c2,c3,c4,c5 in cells:
    print("{} {} {} {} {}".format(c1.value,c2.value,c3.value,c4.value,c5.value))"""

"""from openpyxl import Workbook
wb_file=Workbook()
sheet.merge_cells("F1:N1")
wb_file.save(filename="test.xlsx")"""