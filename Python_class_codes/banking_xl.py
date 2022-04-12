import openpyxl
path = ("/Users/sakethmaddineni/PycharmProjects/project_001/bank_marketing.xlsx")
wb=openpyxl.load_workbook(path)
sheet=wb.active
val1=sheet.cell(row=4,column=5)
val=val1.value
print(val)

for i in range(1,51):
    for j in range(1,6):
        print(sheet.cell(row=i,column=j).value)
