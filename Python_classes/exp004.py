from openpyxl import Workbook

bikes=Workbook()
sheet=bikes.active

class bike:
    def __init__(self,name,price,speed,milage,color,yom):
        self.name=name
        self.price=price
        self.speed=speed
        self.milage = milage
        self.color = color
        self.yom= yom
num=int(input("Enter no.of bikes in the list :"))

sheet['A1'] = "Name"
sheet['B1'] = "Price"
sheet['C1'] = "Speed"
sheet['D1'] = "Milage"
sheet['E1'] = "Color"
sheet['F1'] = "Year of Manufacture"
d=1
i =1
j=1
for d in range(num):
    name=input("Enter the name of the bike :")
    price=input("Enter the price of the bike :")
    speed=input("Enter the top speed of the bike :")
    milage=input("Enter the milage of the bike :")
    color=input("Enter the color of the bike :")
    yom=input("Enter the year of manufacture of the bike :")
    bk = bike(name, price, speed, milage, color, yom)
    d=d+1
    for j in range(num):
        sheet.cell(row=j + 2, column=i).value = bk.name
        sheet.cell(row=j + 2, column=i + 1).value = bk.price
        sheet.cell(row=j + 2, column=i + 2).value = bk.speed
        sheet.cell(row=j + 2, column=i + 3).value = bk.milage
        sheet.cell(row=j + 2, column=i + 4).value = bk.color
        sheet.cell(row=j + 2, column=i + 5).value = bk.yom
        j = j + 1


bikes.save(filename="/Users/sakethmaddineni/Documents/Bike.xlsx")

print("EOF!")
