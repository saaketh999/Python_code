import openpyxl
path=("/Users/sakethmaddineni/PycharmProjects/project_001/Student_Grade_exp3.xlsx")
#Getting name and overall grade of student
wb=openpyxl.load_workbook(path)
sheet=wb.active
val=sheet.cell(row=2,column=2).value
#print(val)
#print(sheet.max_row,sheet.max_column)

#Returning the values in form of tupples
#for i in range(1,sheet.max_row):
 #   for j in range(1,3):
  #      print(sheet.cell(row=i,column=j).value)

#for value in sheet.iter_rows(min_row=1,max_row=sheet.max_row,min_col=1,max_col=2,values_only=True):
 #   print(f"{value[0]}  {value[1]}")
print("Saketh's overall grade")
for value in sheet.iter_rows(min_row=1,max_row=2,min_col=1,max_col=sheet.max_column,values_only=True):
    print(f"{value[0]}  {value[1]}")

#for value1 in sheet.iter_cols(min_row=1,max_row=2,min_col=1,max_col=sheet.max_column,values_only=True):
 #   print(f"{value1}")


